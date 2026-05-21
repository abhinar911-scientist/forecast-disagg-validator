"""
Excel exporter — v2.

Builds a multi-sheet xlsx from a ValidationResult. Snapshot version
(values, not live formulas) — suitable for distribution.
"""
from io import BytesIO
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.worksheet.table import Table, TableStyleInfo

from validator_engine import FLAG_CODES, FLAG_COLORS

# Styles
NAVY = '1F3864'
HEADER_FILL = '1F3864'
WHITE = 'FFFFFF'

thin = Side(border_style='thin', color='B4B4B4')
medium = Side(border_style='medium', color=NAVY)
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
border_header = Border(left=medium, right=medium, top=medium, bottom=medium)

font_title = Font(name='Calibri', size=18, bold=True, color=NAVY)
font_subtitle = Font(name='Calibri', size=11, italic=True, color='595959')
font_section = Font(name='Calibri', size=13, bold=True, color=NAVY)
font_header = Font(name='Calibri', size=10, bold=True, color=WHITE)
font_subheader = Font(name='Calibri', size=10, bold=True, color=NAVY)
font_body = Font(name='Calibri', size=10)
font_kpi_label = Font(name='Calibri', size=10, bold=True, color='595959')
font_kpi_value = Font(name='Calibri', size=22, bold=True, color=NAVY)

fill_header = PatternFill('solid', start_color=HEADER_FILL)
fill_kpi = PatternFill('solid', start_color='F8F9FA')


def flag_fill(code):
    color = FLAG_COLORS.get(code, '#000000').lstrip('#')
    return PatternFill('solid', start_color=color)


def flag_font(code):
    text_color = 'FFFFFF' if code == 'Lost Forecast' else '000000'
    return Font(name='Calibri', size=10, bold=True, color=text_color)


center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_a = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_a = Alignment(horizontal='right', vertical='center')


def _write_header(ws, headers, row, start_col=1):
    for i, h in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = center
        cell.border = border_header


def _add_flag_cf(ws, flag_range):
    """Add conditional formatting that colors each flag value distinctly."""
    for code in FLAG_CODES:
        ws.conditional_formatting.add(
            flag_range,
            CellIsRule(operator='equal', formula=[f'"{code}"'],
                       fill=flag_fill(code), font=flag_font(code))
        )


def _build_summary_sheet(ws, result, settings):
    ws.sheet_view.showGridLines = False
    meta = result.meta

    ws['B2'] = 'Forecast Disaggregation Validation — Summary'
    ws['B2'].font = font_title
    ws.merge_cells('B2:K2')
    rw_start = pd.Timestamp(meta['recent_window_start'])
    rw_end = pd.Timestamp(meta['recent_window_end'])
    fw_start = pd.Timestamp(meta['fcst_window_start'])
    fw_end = pd.Timestamp(meta['fcst_window_end'])
    rw_months = (rw_end.year - rw_start.year) * 12 + (rw_end.month - rw_start.month) + 1
    fw_months = (fw_end.year - fw_start.year) * 12 + (fw_end.month - fw_start.month) + 1
    ws['B3'] = (
        f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')} • "
        f"Recent: {rw_start.strftime('%b %Y')}–{rw_end.strftime('%b %Y')} "
        f"({rw_months} mo) • "
        f"Forecast: {fw_start.strftime('%b %Y')}–{fw_end.strftime('%b %Y')} "
        f"({fw_months} mo) • "
        f"Threshold: ±{settings.threshold:.0%}"
    )
    ws['B3'].font = font_subtitle
    ws.merge_cells('B3:K3')

    # KPIs row (rollup-level)
    rollup_counts = meta['rollup_flag_counts']
    n_total = meta['n_rollup_combos']
    kpis = [
        ('% OK', meta['pct_ok_rollup'], '0.0%'),
        ('Mix Drift', rollup_counts.get('Mix Drift', 0), '#,##0'),
        ('Lost Forecast', rollup_counts.get('Lost Forecast', 0), '#,##0'),
        ('New Demand', rollup_counts.get('New Demand', 0), '#,##0'),
        ('No Activity', rollup_counts.get('No Activity', 0), '#,##0'),
    ]
    starts = ['B', 'E', 'H', 'K', 'N']
    for i, (label, val, fmt) in enumerate(kpis):
        col = starts[i]; nxt = chr(ord(col) + 1)
        ws[f'{col}5'] = label
        ws[f'{col}5'].font = font_kpi_label
        ws[f'{col}5'].alignment = center; ws[f'{col}5'].fill = fill_kpi
        ws.merge_cells(f'{col}5:{nxt}5')
        ws[f'{col}6'] = val
        ws[f'{col}6'].font = font_kpi_value
        ws[f'{col}6'].alignment = center; ws[f'{col}6'].fill = fill_kpi
        ws[f'{col}6'].number_format = fmt
        ws.merge_cells(f'{col}6:{nxt}6')
        for rr in (5, 6):
            for c_l in (col, nxt):
                ws[f'{c_l}{rr}'].border = Border(
                    left=Side('medium', color=NAVY), right=Side('medium', color=NAVY),
                    top=Side('medium', color=NAVY), bottom=Side('medium', color=NAVY))
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 38

    ws.cell(row=8, column=2, value='Flag Distribution').font = font_section
    ws.cell(row=10, column=2, value='Flag').font = font_subheader
    ws.cell(row=10, column=3, value='# Combos').font = font_subheader
    ws.cell(row=10, column=4, value='% of total').font = font_subheader
    for c in (2, 3, 4):
        ws.cell(row=10, column=c).border = border_thin
    for i, code in enumerate(FLAG_CODES):
        rr = 11 + i
        cnt = rollup_counts.get(code, 0)
        ws.cell(row=rr, column=2, value=code)
        ws.cell(row=rr, column=2).fill = flag_fill(code)
        ws.cell(row=rr, column=2).font = flag_font(code)
        ws.cell(row=rr, column=2).alignment = center
        ws.cell(row=rr, column=3, value=cnt)
        ws.cell(row=rr, column=3).number_format = '#,##0'
        ws.cell(row=rr, column=4, value=cnt / n_total if n_total else 0)
        ws.cell(row=rr, column=4).number_format = '0.0%'
        for c in (2, 3, 4):
            ws.cell(row=rr, column=c).border = border_thin

    bar = BarChart()
    bar.type = 'bar'; bar.style = 10
    bar.title = 'Mat × Cust by Flag'
    cats = Reference(ws, min_col=2, min_row=11, max_row=10 + len(FLAG_CODES))
    data = Reference(ws, min_col=3, min_row=10, max_row=10 + len(FLAG_CODES))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 8; bar.width = 14
    ws.add_chart(bar, 'F8')

    rl_violators = result.matcust_rollup_df[result.matcust_rollup_df['Flag'] != 'OK'].copy()
    top20 = rl_violators.nlargest(20, 'Kg Impact')

    section_row = 30
    ws.cell(row=section_row, column=2,
            value='Top 20 Material × Parent Cust Violators (by Kg Impact)').font = font_section
    ws.merge_cells(start_row=section_row, start_column=2, end_row=section_row, end_column=11)

    headers = ['Material', 'Parent Cust', 'Recent (Kg)', 'Forecast (Kg)',
               'Recent Mix %', 'Forecast Mix %', 'Mix Deviation', 'Kg Impact', 'Flag']
    _write_header(ws, headers, row=section_row + 1, start_col=2)
    for i, (_, row) in enumerate(top20.iterrows()):
        rr = section_row + 2 + i
        ws.cell(row=rr, column=2, value=row['Material']).font = font_body
        ws.cell(row=rr, column=3, value=row['Parent Cust']).font = font_body
        ws.cell(row=rr, column=4, value=float(row['Recent Hist (Kg)']))
        ws.cell(row=rr, column=4).number_format = '#,##0'
        ws.cell(row=rr, column=5, value=float(row['Forecast (Kg)']))
        ws.cell(row=rr, column=5).number_format = '#,##0'
        ws.cell(row=rr, column=6, value=float(row['Recent Mix %']))
        ws.cell(row=rr, column=6).number_format = '0.0%'
        ws.cell(row=rr, column=7, value=float(row['Forecast Mix %']))
        ws.cell(row=rr, column=7).number_format = '0.0%'
        ws.cell(row=rr, column=8, value=float(row['Mix Deviation']))
        ws.cell(row=rr, column=8).number_format = '+0.0%;-0.0%;0.0%'
        ws.cell(row=rr, column=9, value=float(row['Kg Impact']))
        ws.cell(row=rr, column=9).number_format = '#,##0'
        ws.cell(row=rr, column=10, value=row['Flag'])
        for c in range(2, 11):
            ws.cell(row=rr, column=c).border = border_thin
            if c in (2, 3, 10):
                ws.cell(row=rr, column=c).alignment = left_a if c <= 3 else center
            else:
                ws.cell(row=rr, column=c).alignment = right_a
            if c != 10:
                ws.cell(row=rr, column=c).font = font_body

    if len(top20) > 0:
        flag_range = f'J{section_row + 2}:J{section_row + 1 + len(top20)}'
        _add_flag_cf(ws, flag_range)

    widths = {'A': 2, 'B': 35, 'C': 28, 'D': 14, 'E': 14, 'F': 14,
              'G': 14, 'H': 14, 'I': 14, 'J': 14, 'K': 14, 'L': 14,
              'M': 14, 'N': 14, 'O': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _build_validation_detail_sheet(ws, df):
    ws.sheet_view.showGridLines = False
    ws['B2'] = 'Validation Detail — Material × Ship To Sub Region × Parent Cust'
    ws['B2'].font = font_title
    ws.merge_cells('B2:N2')

    headers = list(df.columns)
    _write_header(ws, headers, row=4, start_col=2)

    for r_idx, (_, row) in enumerate(df.iterrows(), start=5):
        for c_idx, col in enumerate(headers, start=2):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[col])
            cell.font = font_body
            cell.border = border_thin
            if col in ('Recent Hist (Kg)', 'Forecast (Kg)',
                       'Parent Recent (Kg)', 'Parent Forecast (Kg)',
                       'Kg Impact'):
                cell.number_format = '#,##0'
                cell.alignment = right_a
            elif col in ('Recent Mix %', 'Forecast Mix %'):
                cell.number_format = '0.00%'
                cell.alignment = right_a
            elif col == 'Mix Deviation':
                cell.number_format = '+0.00%;-0.00%;0.00%'
                cell.alignment = right_a
            elif col == '# Details':
                cell.number_format = '0'
                cell.alignment = center
            else:
                cell.alignment = center if col in ('Flag',) else left_a

    last_row = 4 + len(df)
    last_col_letter = get_column_letter(1 + len(headers))
    flag_idx = headers.index('Flag') + 2
    flag_letter = get_column_letter(flag_idx)
    flag_range = f'{flag_letter}5:{flag_letter}{last_row}'
    _add_flag_cf(ws, flag_range)

    dev_idx = headers.index('Mix Deviation') + 2
    dev_letter = get_column_letter(dev_idx)
    dev_range = f'{dev_letter}5:{dev_letter}{last_row}'
    ws.conditional_formatting.add(dev_range,
        ColorScaleRule(start_type='num', start_value=-0.30, start_color='F8696B',
                       mid_type='num', mid_value=0, mid_color='FFFFFF',
                       end_type='num', end_value=0.30, end_color='F8696B'))

    ws.auto_filter.ref = f'B4:{last_col_letter}{last_row}'

    col_widths = {
        'Business Line': 18, 'Material': 32, 'Ship To Sub Region': 16,
        'Parent Cust': 26, '# Details': 10,
        'Recent Hist (Kg)': 14, 'Forecast (Kg)': 14,
        'Parent Recent (Kg)': 16, 'Parent Forecast (Kg)': 16,
        'Recent Mix %': 12, 'Forecast Mix %': 14,
        'Mix Deviation': 13, 'Kg Impact': 13, 'Flag': 14,
    }
    for i, col in enumerate(headers, start=2):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 14)
    ws.column_dimensions['A'].width = 2
    ws.freeze_panes = 'B5'


def _build_rollup_sheet(ws, df):
    ws.sheet_view.showGridLines = False
    ws['B2'] = 'Material × Parent Customer Rollup'
    ws['B2'].font = font_title
    ws.merge_cells('B2:M2')

    headers = list(df.columns)
    _write_header(ws, headers, row=4, start_col=2)

    for r_idx, (_, row) in enumerate(df.iterrows(), start=5):
        for c_idx, col in enumerate(headers, start=2):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[col])
            cell.font = font_body
            cell.border = border_thin
            if col in ('Recent Hist (Kg)', 'Forecast (Kg)',
                       'Parent Recent (Kg)', 'Parent Forecast (Kg)',
                       'Kg Impact'):
                cell.number_format = '#,##0'
                cell.alignment = right_a
            elif col in ('Recent Mix %', 'Forecast Mix %'):
                cell.number_format = '0.00%'
                cell.alignment = right_a
            elif col == 'Mix Deviation':
                cell.number_format = '+0.00%;-0.00%;0.00%'
                cell.alignment = right_a
            elif col in ('# SubRegions',):
                cell.number_format = '0'
                cell.alignment = center
            else:
                cell.alignment = center if col == 'Flag' else left_a

    last_row = 4 + len(df)
    flag_idx = headers.index('Flag') + 2
    flag_letter = get_column_letter(flag_idx)
    flag_range = f'{flag_letter}5:{flag_letter}{last_row}'
    _add_flag_cf(ws, flag_range)

    dev_idx = headers.index('Mix Deviation') + 2
    dev_letter = get_column_letter(dev_idx)
    dev_range = f'{dev_letter}5:{dev_letter}{last_row}'
    ws.conditional_formatting.add(dev_range,
        ColorScaleRule(start_type='num', start_value=-0.30, start_color='F8696B',
                       mid_type='num', mid_value=0, mid_color='FFFFFF',
                       end_type='num', end_value=0.30, end_color='F8696B'))

    last_col_letter = get_column_letter(1 + len(headers))
    ws.auto_filter.ref = f'B4:{last_col_letter}{last_row}'

    col_widths = {
        'Material': 32, 'Parent Cust': 26, '# SubRegions': 12,
        'Recent Hist (Kg)': 14, 'Forecast (Kg)': 14,
        'Parent Recent (Kg)': 16, 'Parent Forecast (Kg)': 16,
        'Recent Mix %': 12, 'Forecast Mix %': 14,
        'Mix Deviation': 13, 'Kg Impact': 13, 'Flag': 14,
    }
    for i, col in enumerate(headers, start=2):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 14)
    ws.column_dimensions['A'].width = 2
    ws.freeze_panes = 'B5'


def _build_settings_sheet(ws, settings, meta):
    ws.sheet_view.showGridLines = False
    ws['B2'] = 'Settings & Methodology'
    ws['B2'].font = font_title
    ws.merge_cells('B2:E2')

    rw_start = pd.Timestamp(meta['recent_window_start'])
    rw_end = pd.Timestamp(meta['recent_window_end'])
    fw_start = pd.Timestamp(meta['fcst_window_start'])
    fw_end = pd.Timestamp(meta['fcst_window_end'])

    rows = [
        ('Recent History Window',
         f"{rw_start.strftime('%b %Y')} to {rw_end.strftime('%b %Y')}", ''),
        ('Forecast Horizon',
         f"{fw_start.strftime('%b %Y')} to {fw_end.strftime('%b %Y')}", ''),
        ('Deviation Threshold', f"±{settings.threshold:.1%}", ''),
        ('', '', ''),
        ('Detected Date Anchors (from data)', '', ''),
        ('Last history month',
         meta['last_hist_date'].strftime('%Y-%m-%d'), ''),
        ('First forecast month',
         meta['first_fcst_date'].strftime('%Y-%m-%d'), ''),
        ('Last forecast month',
         meta['last_fcst_date'].strftime('%Y-%m-%d'), ''),
        ('', '', ''),
        ('Flag Reason Codes', '', ''),
        ('OK', 'Both > 0, |deviation| within threshold', ''),
        ('Mix Drift', 'Both > 0 but |deviation| exceeds threshold', ''),
        ('New Demand', 'History = 0, Forecast > 0', ''),
        ('Lost Forecast', 'History > 0, Forecast = 0', ''),
        ('No Activity', 'Both = 0 (dormant)', ''),
    ]
    _write_header(ws, ['Parameter', 'Value', ''], row=4, start_col=2)
    for i, (param, val, _) in enumerate(rows):
        rr = 5 + i
        if param and val:
            ws.cell(row=rr, column=2, value=param).font = font_body
            ws.cell(row=rr, column=3, value=val).font = font_body
            ws.cell(row=rr, column=2).border = border_thin
            ws.cell(row=rr, column=3).border = border_thin
            ws.cell(row=rr, column=2).alignment = left_a
            ws.cell(row=rr, column=3).alignment = left_a
        elif param:
            cell = ws.cell(row=rr, column=2, value=param)
            cell.font = font_section

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 50


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
def build_excel_export(result, settings) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('Summary')
    _build_summary_sheet(ws, result, settings)

    ws = wb.create_sheet('Validation Detail')
    _build_validation_detail_sheet(ws, result.validation_df)

    ws = wb.create_sheet('MatCust Rollup')
    _build_rollup_sheet(ws, result.matcust_rollup_df)

    ws = wb.create_sheet('Settings')
    _build_settings_sheet(ws, settings, result.meta)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
